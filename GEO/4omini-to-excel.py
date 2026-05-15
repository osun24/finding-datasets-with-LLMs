import json
import pandas as pd
import re
from datetime import datetime
from glob import glob
from pathlib import Path
from typing import List

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

""""4.1-mini/4.1mini-chemo-output.jsonl","4.1-mini/4.1mini-chemo-output-v2.jsonl", "4.1-mini/4.1mini-chemo-output-v3.jsonl", 
              "o4-mini/o4mini-chemo-output.jsonl", "o4-mini/o4mini-chemo-output-v2.jsonl", "o4-mini/o4mini-chemo-output-v3.jsonl", "4o-mini/4omini-chemo-output.jsonl", "4o-mini/4omini-chemo-output-v2.jsonl", "4o-mini/4omini-chemo-output-v3.jsonl" """

use_prompt_set = True
SPLIT_NAME = "prompt-set" if use_prompt_set else "test-set"
DISCOVER_SPLITS = ["prompt-set", "test-set"]

PROMPT_POSITIVE_FILES = [
    "samples_table_GSE101929.csv",
    "samples_table_GSE16025.csv",
    "samples_table_GSE37745.csv",
    "samples_table_GSE39279.csv",
]

TEST_POSITIVE_FILES = [
    "samples_table_GSE102287.csv",
    "samples_table_GSE14814.csv",
    "samples_table_GSE19188.csv",
    "samples_table_GSE29013.csv",
    "samples_table_GSE42127.csv",
    "samples_table_GSE42425.csv",
    "samples_table_GSE47115.csv",
    "samples_table_GSE50081.csv",
]

COMBINED_POSITIVE_FILES = PROMPT_POSITIVE_FILES + TEST_POSITIVE_FILES
ACTIVE_POSITIVE_FILES = PROMPT_POSITIVE_FILES if use_prompt_set else TEST_POSITIVE_FILES

# Seed of 42 by default for gpt-5-mini and o4-mini models
BASE_FILE_PATHS = [
    f"out-gpt-5-mini-chemo-may13-v1-prompt-set.jsonl",
    f"out-gpt-4o-mini-chemo-may13-v1-prompt-set.jsonl",
    f"out-gpt-4.1-mini-chemo-may13-v1-prompt-set.jsonl",
    f"out-o4-mini-chemo-may13-v1-prompt-set.jsonl",
    
    f"out-gpt-5-mini-chemo-may13-v2-prompt-set.jsonl",
    f"out-gpt-4o-mini-chemo-may13-v2-prompt-set.jsonl",
    f"out-gpt-4.1-mini-chemo-may13-v2-prompt-set.jsonl",
    f"out-o4-mini-chemo-may13-v2-prompt-set.jsonl",
    
    f"out-gpt-5-mini-chemo-may13-v2-test-set.jsonl",
    f"out-gpt-4o-mini-chemo-may13-v2-test-set.jsonl",
    f"out-gpt-4.1-mini-chemo-may13-v2-test-set.jsonl",
    f"out-o4-mini-chemo-may13-v2-test-set.jsonl",
]

REPLICATION_PATTERNS = [
    f"out-gpt-5-mini-chemo-may13-v*-test-set-*.jsonl",
    f"out-o4-mini-chemo-may13-v*-test-set-*.jsonl",
]


def discover_file_paths() -> List[str]:
    """Gather all batch output files, including replication seeds."""
    discovered = set()
    for path in BASE_FILE_PATHS:
        if Path(path).exists():
            discovered.add(path)
    for pattern in REPLICATION_PATTERNS:
        for path in glob(pattern):
            if Path(path).is_file():
                discovered.add(path)
    return sorted(discovered)


file_paths = discover_file_paths()

MODEL_ALIASES = sorted(
    [
        ("gpt-5-mini", "GPT-5-mini"),
        ("gpt-4o-mini", "GPT-4o-mini"),
        ("4o-mini", "GPT-4o-mini"),
        ("o4-mini", "o4-mini"),
        ("gpt-4.1-mini", "GPT-4.1-mini"),
        ("4.1mini", "GPT-4.1-mini"),
        ("gpt-4.1-nano", "GPT-4.1-nano"),
        ("4.1nano", "GPT-4.1-nano"),
    ],
    key=lambda item: len(item[0]),
    reverse=True,
)

if not file_paths:
    print("⚠️ No batch output files found. Update the path patterns or download batch results first.")

ROOT_CAUSE_OPTIONS = [
    "disease_scope_misread",
    "healthy_controls_misread",
    "control_percentage_misread",
    "proteomics_evidence_misread",
    "cell_line_vs_patient_confusion",
    "ovarian_vs_ovary_confusion",
    "mechanistic_study_overincluded",
    "rationale_correct_final_answer_wrong",
    "metadata_insufficient",
    "human_label_suspect",
    "other",
]

STRUCTURED_FIELD_MAP = {
    "q1_survival_data": "Q1",
    "q2_stage_i_or_act": "Q2",
    "q3_inclusion_justification": "Q3",
    "q4_include_in_meta_analysis": "Q4",
}


def format_structured_answer(value):
    if isinstance(value, dict):
        parts = []
        if "answer" in value:
            parts.append(str(value["answer"]))
        for key, item in value.items():
            if key == "answer" or item in (None, ""):
                continue
            label = key.replace("_", " ")
            parts.append(f"{label}: {item}")
        return " | ".join(parts)
    if value is None:
        return ""
    return str(value)


def parse_structured_answers(content):
    try:
        parsed = json.loads(content)
    except (TypeError, json.JSONDecodeError) as exc:
        raise ValueError("Expected structured JSON content from model response.") from exc

    if not isinstance(parsed, dict):
        raise ValueError("Expected structured JSON object from model response.")

    answers = {}
    for source_key, target_key in STRUCTURED_FIELD_MAP.items():
        if source_key in parsed:
            answers[target_key] = format_structured_answer(parsed[source_key])

    missing_keys = sorted(set(STRUCTURED_FIELD_MAP) - set(parsed))
    if missing_keys:
        raise ValueError(f"Structured JSON response missing keys: {missing_keys}")

    return answers


def empty_structured_answers():
    return {target_key: "" for target_key in STRUCTURED_FIELD_MAP.values()}


def question_sort_key(col):
    match = re.match(r'^Q(\d+)([a-z]?)$', col, re.IGNORECASE)
    if not match:
        return (0, "")
    return (int(match.group(1)), match.group(2))


def get_question_columns(df):
    return sorted([c for c in df.columns if c.startswith("Q")], key=question_sort_key)


def get_final_answer_column(df):
    q_cols = get_question_columns(df)
    return q_cols[-1] if q_cols else df.columns[-1]


def sheet_name_for(model, version):
    sheet_name = re.sub(r"[\[\]\:\*\?\/\\]", "_", f"{model}_{version}")
    return sheet_name[:31]


def split_for_file(file_path):
    filename = Path(file_path).name
    if "-prompt-set" in filename:
        return "prompt-set", PROMPT_POSITIVE_FILES
    if "-test-set" in filename:
        return "test-set", TEST_POSITIVE_FILES
    raise ValueError(f"Cannot infer split from filename: {filename}")


def prompt_version_number(version):
    match = re.fullmatch(r"v(\d+)", str(version))
    return int(match.group(1)) if match else None


def latest_reviewed_prompt_version(output_path):
    output_path = Path(output_path)
    review_dir = output_path.parent if output_path.parent != Path("") else Path(".")
    latest_version = 0

    for path in review_dir.glob("false_classification*_review_v*.xlsx"):
        match = re.fullmatch(r"false_classifications?_review_v(\d+)\.xlsx", path.name)
        if match:
            latest_version = max(latest_version, int(match.group(1)))

    return latest_version


# Initialize results storage
def load_or_create_results():
    """Load existing results or create new structure"""
    results_file = "model_performance_results.json"
    try:
        with open(results_file, 'r') as f:
            existing_results = json.load(f)
        print(f"📂 Loaded existing results from {results_file}")
        return existing_results
    except FileNotFoundError:
        print(f"📝 Creating new results file: {results_file}")
        return {
            "metadata": {
                "timestamp": datetime.now().isoformat(),
                "description": "AI model performance evaluation for clinical dataset screening",
                "total_actual_studies": 0,
                "evaluation_criteria": "NSCLC clinical dataset inclusion/exclusion",
                "last_updated": datetime.now().isoformat()
            },
            "results": []
        }

all_results = load_or_create_results()

def parse_batch_output(file_path, csv_output=None, actual_files=None):
    actual_files = set(actual_files or [])
    rows = []
    for line in open(file_path, 'r'):
        data = json.loads(line)
        response = data.get('response') or {}
        body = response.get('body') or {}
        choices = body.get('choices') or []
        message = choices[0].get('message', {}) if choices else {}
        content = message.get('content') or ''
        model = body.get('model', '')
        parse_error = ""
        try:
            answers = parse_structured_answers(content)
        except ValueError as exc:
            answers = empty_structured_answers()
            finish_reason = choices[0].get("finish_reason", "") if choices else ""
            parse_error = f"parse_error: {exc}"
            if finish_reason:
                parse_error = f"{parse_error} (finish_reason={finish_reason})"
        
        cid = data.get('custom_id', '')
        error = data.get('error', '') or message.get('refusal', '')
        if parse_error:
            error = f"{error}; {parse_error}" if error else parse_error
        row = {
            "custom_id": cid,
            "error": error,
            "model": model,
        }
        row.update(answers)
        row['in_actual'] = cid in actual_files
        rows.append(row)

    df = pd.DataFrame(rows)
    if csv_output:
        df.to_csv(csv_output, index=False)
        print(f"Parsed batch output saved to {csv_output}")
    return df

""" MATCH """
actual = ACTIVE_POSITIVE_FILES

def match_files(user_input, actual):
    matched = []
    false_positives = []  # User input but not in actual
    false_negatives = []  # In actual but not in user input
    
    for item in user_input:
        if item in actual:
            matched.append(item)
        else:
            false_positives.append(item)
    
    for item in actual:
        if item not in user_input:
            false_negatives.append(item)
    
    return matched, false_positives, false_negatives  

# calculate sensitivity, specificity, and accuracy
def calculate_metrics(matched, false_positives, false_negatives, total_files_evaluated):
    true_positive = len(matched)
    false_positive = len(false_positives)
    false_negative = len(false_negatives)
    
    # True Negatives: files that were correctly identified as NOT meeting criteria
    # This is total files evaluated minus all the positive cases (TP + FP + FN)
    true_negative = total_files_evaluated - (true_positive + false_positive + false_negative)
    
    print(f"True Positive: {true_positive}")
    print(f"False Positive: {false_positive}")
    print(f"False Negative: {false_negative}")
    print(f"True Negative: {true_negative}")
    print(f"Total files evaluated: {total_files_evaluated}")
    
    sensitivity = true_positive / (true_positive + false_negative) if (true_positive + false_negative) else 0
    specificity = true_negative / (true_negative + false_positive) if (true_negative + false_positive) else 0
    
    # Correct accuracy calculation: (TP + TN) / (TP + FP + FN + TN)
    accuracy = (true_positive + true_negative) / total_files_evaluated if total_files_evaluated else 0
    
    # Precision and F1 score
    precision = true_positive / (true_positive + false_positive) if (true_positive + false_positive) else 0
    f1_score = 2 * (precision * sensitivity) / (precision + sensitivity) if (precision + sensitivity) else 0
    
    print(f"Sensitivity (Recall): {sensitivity:.3f}")
    print(f"Specificity: {specificity:.3f}")
    print(f"Precision: {precision:.3f}")
    print(f"Accuracy: {accuracy:.3f}")
    print(f"F1 Score: {f1_score:.3f}")
    
    # Return all metrics as a dictionary for JSON storage
    return {
        "true_positive": true_positive,
        "false_positive": false_positive,
        "false_negative": false_negative,
        "true_negative": true_negative,
        "total_files_evaluated": total_files_evaluated,
        "sensitivity": round(sensitivity, 3),
        "specificity": round(specificity, 3),
        "precision": round(precision, 3),
        "accuracy": round(accuracy, 3),
        "f1_score": round(f1_score, 3)
    }

def extract_model_metadata(file_path):
    """Extract model, prompt version, and trial identifier from a batch output file."""
    filename = Path(file_path).name

    model = "Unknown"
    for pattern, canonical in MODEL_ALIASES:
        if pattern in filename:
            model = canonical
            break

    version_match = re.search(r'v(\d+)', filename)
    version = f"v{version_match.group(1)}" if version_match else "v1"

    seed_match = re.search(r'seed(\d+)', filename)
    if seed_match:
        trial = f"seed{seed_match.group(1)}"
    else:
        alt_seed_match = re.search(
            r'-v\d+(?:-(?:prompt|test)-set)?-([0-9]+)(?=\.jsonl$)',
            filename,
        )
        trial = f"seed{alt_seed_match.group(1)}" if alt_seed_match else None

    return {
        "model": model,
        "version": version,
        "trial": trial,
    }

def save_results_to_json(all_results, filename="model_performance_results.json"):
    """Save all results to JSON file"""
    with open(filename, 'w') as f:
        json.dump(all_results, indent=2, fp=f)
    print(f"\n📊 Results saved to {filename}")
    return filename

def show_detailed_responses(df, false_positives, false_negatives, first_col):
    """Show detailed AI responses for False Positives and False Negatives"""
    
    if false_positives:
        print("\n" + "="*60)
        print("FALSE POSITIVES - Detailed Responses:")
        print("="*60)
        for fp_file in false_positives:
            fp_row = df[df[first_col] == fp_file]
            if not fp_row.empty:
                print(f"\n📁 {fp_file}")
                print("-" * 50)
                # Show all Q columns
                for col in df.columns:
                    if col.startswith('Q') and not fp_row[col].isna().iloc[0]:
                        q_num = col[1:]  # Remove 'Q' prefix
                        response = fp_row[col].iloc[0]
                        print(f"Q{q_num}: {response}")
                print()
    
    if false_negatives:
        print("\n" + "="*60)
        print("FALSE NEGATIVES - Detailed Responses:")
        print("="*60)
        for fn_file in false_negatives:
            fn_row = df[df[first_col] == fn_file]
            if not fn_row.empty:
                print(f"\n📁 {fn_file}")
                print("-" * 50)
                # Show all Q columns
                for col in df.columns:
                    if col.startswith('Q') and not fn_row[col].isna().iloc[0]:
                        q_num = col[1:]  # Remove 'Q' prefix
                        response = fn_row[col].iloc[0]
                        print(f"Q{q_num}: {response}")
                print()
            else:
                print(f"\n📁 {fn_file} - NO AI RESPONSE FOUND (not in evaluated dataset)")
                print("-" * 50)


def build_error_review_dataframe(df, model, version, file_path, split_label):
    """Build review rows for false positives and false negatives only."""
    if df.empty:
        return pd.DataFrame()

    first_col = df.columns[0]
    q_cols = get_question_columns(df)
    final_col = get_final_answer_column(df)
    model_include = df[final_col].astype(str).str.contains("yes", case=False, na=False)
    human_include = df["in_actual"].astype(bool)
    error_mask = model_include.ne(human_include)

    if not error_mask.any():
        return pd.DataFrame()

    review_df = pd.DataFrame(
        {
            "dataset_id": df.loc[error_mask, first_col],
            "source_file": str(file_path),
            "model": model,
            "version": version,
            "split": split_label,
            "human_include": human_include.loc[error_mask].values,
            "model_include": model_include.loc[error_mask].values,
            "error_type": model_include.loc[error_mask].map(lambda value: "FP" if value else "FN").values,
        }
    )

    for col in q_cols:
        review_df[col] = df.loc[error_mask, col].values

    review_df["root_cause"] = ""
    review_df["notes"] = ""
    return review_df


def autosize_review_columns(worksheet):
    widths = {
        "dataset_id": 28,
        "source_file": 44,
        "model": 16,
        "version": 10,
        "split": 18,
        "human_include": 14,
        "model_include": 14,
        "error_type": 12,
        "root_cause": 32,
        "notes": 42,
    }
    for cell in worksheet[1]:
        column_name = str(cell.value)
        column_letter = cell.column_letter
        if column_name.startswith("Q"):
            worksheet.column_dimensions[column_letter].width = 58
        else:
            worksheet.column_dimensions[column_letter].width = widths.get(column_name, 18)


def format_review_sheet(worksheet, root_cause_formula, row_count, col_count):
    if row_count < 1 or col_count < 1:
        return

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    fp_fill = PatternFill("solid", fgColor="FCE4D6")
    fn_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    cell_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment
        cell.border = cell_border

    for row in worksheet.iter_rows(min_row=2, max_row=row_count, max_col=col_count):
        for cell in row:
            cell.alignment = wrap_alignment
            cell.border = cell_border

    worksheet.freeze_panes = "B2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_review_columns(worksheet)

    headers = [cell.value for cell in worksheet[1]]
    error_type_col = headers.index("error_type") + 1 if "error_type" in headers else None
    root_cause_col = headers.index("root_cause") + 1 if "root_cause" in headers else None

    if error_type_col:
        error_col_letter = worksheet.cell(row=1, column=error_type_col).column_letter
        data_range = f"A2:{worksheet.cell(row=row_count, column=col_count).coordinate}"
        worksheet.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'${error_col_letter}2="FP"'], fill=fp_fill),
        )
        worksheet.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'${error_col_letter}2="FN"'], fill=fn_fill),
        )

    if root_cause_col and row_count >= 2:
        root_cause_letter = worksheet.cell(row=1, column=root_cause_col).column_letter
        validation = DataValidation(
            type="list",
            formula1=root_cause_formula,
            allow_blank=True,
        )
        worksheet.add_data_validation(validation)
        validation.add(f"{root_cause_letter}2:{root_cause_letter}{row_count}")


def save_error_review_workbook(review_frames, output_path="false_classifications_review.xlsx"):
    populated_frames = [frame for frame in review_frames if not frame.empty]
    if not populated_frames:
        print("No false classifications found; review workbook was not created.")
        return None

    latest_reviewed_version = latest_reviewed_prompt_version(output_path)
    if latest_reviewed_version:
        populated_frames = [
            frame
            for frame in populated_frames
            if prompt_version_number(frame["version"].iloc[0]) is None
            or prompt_version_number(frame["version"].iloc[0]) > latest_reviewed_version
        ]

    if not populated_frames:
        print(
            "No false classifications found for prompt versions after "
            f"v{latest_reviewed_version}; review workbook was not created."
        )
        return None

    grouped_frames = {}
    for frame in populated_frames:
        model = frame["model"].iloc[0]
        version = frame["version"].iloc[0]
        sheet_name = sheet_name_for(model, version)
        grouped_frames.setdefault(sheet_name, []).append(frame)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        root_causes_df = pd.DataFrame({"root_cause": ROOT_CAUSE_OPTIONS})
        root_causes_df.to_excel(writer, sheet_name="_root_causes", index=False)
        root_causes_sheet = writer.sheets["_root_causes"]
        root_causes_sheet.sheet_state = "hidden"
        root_cause_formula = f"'_root_causes'!$A$2:$A${len(ROOT_CAUSE_OPTIONS) + 1}"

        for sheet_name, frames in sorted(grouped_frames.items()):
            sheet_df = pd.concat(frames, ignore_index=True)
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            format_review_sheet(
                worksheet,
                root_cause_formula=root_cause_formula,
                row_count=len(sheet_df) + 1,
                col_count=len(sheet_df.columns),
            )

        first_visible_sheet = next(
            index
            for index, worksheet in enumerate(writer.book.worksheets)
            if worksheet.sheet_state == "visible"
        )
        writer.book.active = first_visible_sheet

    print(f"False-classification review workbook saved to {output_path}")
    return output_path


def evaluate_dataframe(df, model, version, trial, file_path, actual_files, split_label):
    # Determine the last question column (columns named like 'Q1', 'Q2', ...).
    # We can't assume it's the final DataFrame column because we add fields
    # like `in_actual` which may come after the Q columns.
    last_col = get_final_answer_column(df)

    # Sort by the last question column in ascending order
    df_sorted = df.sort_values(by=last_col, ascending=True)

    # Filter rows where the last column contains "yes" (case-insensitive)
    filtered = df_sorted[df_sorted[last_col].str.contains("yes", case=False, na=False)]

    print(f"Identified studies from {last_col}: {len(filtered)}")

    # print identified studies by printing first column in filtered
    first_col = df.columns[0]
    identified_studies = filtered[first_col].tolist()

    # Run through matching process
    matched, false_positives, false_negatives = match_files(identified_studies, actual_files)

    if matched:
        print("> Matched files:")
        for item in matched:
            print("  " + item)
    else:
        print("No files matched.")

    if false_positives:
        print("\n> False Positives (FP):")
        for item in false_positives:
            print(f"  {item}")

    if false_negatives:
        print("\n> False Negatives (FN):")
        for item in false_negatives:
            print(f"  {item}")

    # Calculate and display metrics
    total_files_evaluated = len(df)  # Total number of files/studies evaluated by the AI
    metrics = calculate_metrics(matched, false_positives, false_negatives, total_files_evaluated)

    result_entry = {
        "model": model,
        "version": version,
        "trial": trial,
        "split": split_label,
        "file_path": file_path,
        "timestamp": datetime.now().isoformat(),
        "metrics": metrics,
        "details": {
            "identified_studies": identified_studies,
            "matched_studies": matched,
            "false_positives": false_positives,
            "false_negatives": false_negatives,
            "total_studies_in_dataset": len(actual_files),
            "trial": trial,
            "split": split_label,
        }
    }

    # Show detailed responses for FP and FN
    show_detailed_responses(df, false_positives, false_negatives, first_col)
    return result_entry


def upsert_result(result_entry):
    existing_index = None
    for i, existing_result in enumerate(all_results["results"]):
        if (
            existing_result.get("model") == result_entry["model"]
            and existing_result.get("version") == result_entry["version"]
            and existing_result.get("file_path") == result_entry["file_path"]
            and (existing_result.get("trial") or "") == (result_entry.get("trial") or "")
            and (existing_result.get("split") or "") == (result_entry.get("split") or "")
        ):
            existing_index = i
            break

    if existing_index is not None:
        print(
            f"Updating existing entry for {result_entry['model']} "
            f"{result_entry['version']} [{result_entry['split']}]"
        )
        all_results["results"][existing_index] = result_entry
    else:
        print(
            f"Adding new entry for {result_entry['model']} "
            f"{result_entry['version']} [{result_entry['split']}]"
        )
        all_results["results"].append(result_entry)

    all_results["metadata"]["last_updated"] = datetime.now().isoformat()


def prompt_pair_for_test_file(file_path):
    prompt_path = Path(str(file_path).replace("-test-set", "-prompt-set"))
    return str(prompt_path) if prompt_path.is_file() else None
    
# Set the total actual studies count
discovered_splits = sorted({split_for_file(file_path)[0] for file_path in file_paths})
if discovered_splits == ["prompt-set"]:
    total_actual_studies = len(PROMPT_POSITIVE_FILES)
elif discovered_splits == ["test-set"]:
    total_actual_studies = len(TEST_POSITIVE_FILES)
else:
    total_actual_studies = len(COMBINED_POSITIVE_FILES)

all_results["metadata"]["total_actual_studies"] = total_actual_studies
all_results["metadata"]["prompt_set_actual_studies"] = len(PROMPT_POSITIVE_FILES)
all_results["metadata"]["test_set_actual_studies"] = len(TEST_POSITIVE_FILES)
all_results["metadata"]["combined_actual_studies"] = len(COMBINED_POSITIVE_FILES)
all_results["metadata"]["split"] = ",".join(discovered_splits) if discovered_splits else SPLIT_NAME
error_review_frames = []

for file_path in file_paths:
    csv_output = file_path.replace('.jsonl', '.csv')
    metadata = extract_model_metadata(file_path)
    model = metadata["model"]
    version = metadata["version"]
    trial = metadata["trial"]
    split_label, actual_files = split_for_file(file_path)

    label = f"{model} {version}".strip()
    if trial:
        label = f"{label} [{trial}]"

    print(f"Processing {file_path} → {label} [{split_label}]...")
    df = parse_batch_output(file_path, csv_output, actual_files=actual_files)
    result_entry = evaluate_dataframe(
        df=df,
        model=model,
        version=version,
        trial=trial,
        file_path=file_path,
        actual_files=actual_files,
        split_label=split_label,
    )
    upsert_result(result_entry)
    if split_label == "prompt-set":
        error_review_frames.append(
            build_error_review_dataframe(
                df=df,
                model=model,
                version=version,
                file_path=file_path,
                split_label=split_label,
            )
        )

    if split_label == "test-set":
        prompt_file_path = prompt_pair_for_test_file(file_path)
        if prompt_file_path:
            print(f"\nProcessing combined prompt+test set for {label}...")
            prompt_df = parse_batch_output(
                prompt_file_path,
                actual_files=PROMPT_POSITIVE_FILES,
            )
            combined_df = pd.concat([prompt_df, df], ignore_index=True)
            combined_file_path = f"combined:{prompt_file_path}+{file_path}"
            combined_entry = evaluate_dataframe(
                df=combined_df,
                model=model,
                version=version,
                trial=trial,
                file_path=combined_file_path,
                actual_files=COMBINED_POSITIVE_FILES,
                split_label="combined-prompt-test",
            )
            upsert_result(combined_entry)
        else:
            print(
                f"No prompt-set pair found for {file_path}; "
                "skipping combined prompt+test calculation."
            )

    print("\n" + "="*40 + "\n")

# Save all results to JSON
save_results_to_json(all_results)
save_error_review_workbook(error_review_frames)

# Also create a simplified DataFrame for quick analysis
summary_data = []
for result in all_results["results"]:
    row = {
        "Model": result.get("model", "Unknown"),
        "Version": result.get("version", ""),
        "Trial": result.get("trial", ""),
        "Split": result.get("split", ""),
        "File": result.get("file_path", ""),
        **result["metrics"]  # Unpack all metrics
    }
    summary_data.append(row)

summary_df = pd.DataFrame(summary_data)

if not summary_df.empty:
    sort_cols = [col for col in ["Model", "Version", "Trial", "Split"] if col in summary_df.columns]
    summary_df = summary_df.sort_values(by=sort_cols).reset_index(drop=True)

summary_csv = "model_performance_summary.csv"
summary_df.to_csv(summary_csv, index=False)
print(f"📈 Summary table saved to {summary_csv}")

# Print final summary
print("\n" + "="*60)
print("FINAL SUMMARY:")
print("="*60)
for result in all_results["results"]:
    metrics = result["metrics"]
    split = result.get("split", "")
    split_suffix = f" [{split}]" if split else ""
    print(f"{result['model']} {result['version']}{split_suffix}: "
          f"Acc={metrics['accuracy']:.3f}, "
          f"Sen={metrics['sensitivity']:.3f}, "
          f"Spe={metrics['specificity']:.3f}, "
          f"F1={metrics['f1_score']:.3f}")
